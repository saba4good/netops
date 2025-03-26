'''
Objectives
- Read in Citrix inspection text files and Citrix excel output files and a template excel file
- Create a folder to place running config files and a report excel file
To do
- Read in files less times

Changes
v. 0.7.1 : 2025.03.21
- Changed how to handle missing CSV files (Expect some files or fields missing and still run)
v. 0.4 : 2025.02.25
- Function parse_citrix_cmd_output() is changed and finished for text output files (except intf errors)
- Function fill_excel_with_nested_data() is done for summary sheet
- Tested on Ubuntu 22.04
v. 0.2 : 2025.02.25
- Function parse_citrix_cmd_output() is finished for text output files
- Tested on Ubuntu 22.04
v. 0.1 : authored by HYP. 2025.02.20
- Tested on Ubuntu 22.04
'''
import logging_setup as Logger
from datetime import datetime
from zoneinfo import ZoneInfo
from dataclasses import dataclass
from pathlib import Path
import re
import csv, json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

# Constants
CPU_THRESHOLD = 50
MEM_THRESHOLD = 70
HA_PRIMARY_COLOR = 'e4fbd5'
WARN_FILL = PatternFill(fgColor='FFAA00', fill_type='solid')

# dataclass cannot have mutable object initialization such as a list or a dictionary
@dataclass
class CitrixData:
    hostname: str = ""
    model: str = ""
    serial: str = ""
    license_type: str = ""
    os_version: str = ""
    uptime: str = ""
    ns_ip: str = ""
    gateway: str = ""
    cpu_util: dict = None     #{'Current':'', 'MAX':''}  # Packet CPU usage (%)
    memory_util: dict = None  #{'Current':'', 'MAX':''}  # InUse Memory (%)
    mgmt_cpu_util: dict = None
    fans: dict = None       #{'rpm': 0}
    temperature: dict = None
    powers: dict = None     #{'health':''}
    intf_errors: str = ""
    ha: dict = None #{'health':'', 'state':{'type':'', 'time':''}, 'sync':'', 'prop':''}
    slb_count: dict = None   # This will have vs_count, svc_count, svr_count dict
    ssl_cards: str = ""
    ssl_certs: dict = None  #{'valid_cert_name':'number of dates until expiration'}
    tcp_cc_client: dict = None  #{'Current':0, 'MAX':0}  # Established client connections
    tcp_cc_server: dict = None  #{'Current':0, 'MAX':0}  # Established server connections
    tcp_cps: dict = None  #{'Current':0, 'MAX':0}
    http_rps: dict = None  #{'Current':0, 'MAX':0}  # Total requests (Rate),Total responses (Rate)
    ssl_sps: dict = None  #{'Current':0, 'MAX':0}   # SSL sessions  (Rate)
    ssl_tps: dict = None  #{'Current':0, 'MAX':0}
    key_eps_rsa2048: dict = None  #{'Current':0, 'MAX':0}  # RSA 2048-bit key exchanges (Rate)
    key_eps_ecdhe256: dict = None  #{'Current':0, 'MAX':0}  # ECDHE 256 curve key exchanges (Rate)
    throughput: dict = None  #{'Current':0, 'MAX':0}  # Megabits received (Rate),Megabits transmitted (Rate)
    log: str = ""
    date_time: str = ""
    log_file_name: str = ""

def initialize_paths():
    """Initialize input and output paths."""
    directory = Path("./")
    input_folder = directory / "log"
    today = datetime.today().strftime("%Y%m%d_%H%M")
    report_folder = Path(f"./out_{today}")

    # Ensure unique folder name
    idx = 0
    while report_folder.exists():
        idx += 1
        report_folder = Path(f"./out_{today}_{idx}")

    report_folder.mkdir(parents=True, exist_ok=True)
    return input_folder, report_folder

def load_and_prepare_input_files(input_folder, report_folder, cmd_csv_file):
    """Process text files in the input folder."""
    commands_dict = {}
    status_data = {}
    for file in input_folder.glob('*.log'):
        #logging.debug(f"Processing file: {file}")
        commands_dict[file] = sanitize_input_text_file(file, report_folder)
        device_data = parse_citrix_cmd_output(file)
        #device_data.cpu_max_util = find_max_cpu_util_from_file(file)
        #device_data.intf_errors = parse_interface_errors(file)
        ##logging.debug(f"{device_data.hostname} interface errors: {device_data.intf_errors}")
        #logging.debug(f"Final data for {device_data.hostname}: {device_data}\n")
        status_data[device_data.hostname] = device_data.__dict__
    # Write raw data into json and csv
    write_device_data_to_json(report_folder, status_data)
    with open(cmd_csv_file, 'w', newline='') as csv_f:
        writer = csv.writer(csv_f)
        # Optionally, write a header:
        writer.writerow(['File', 'Command'])
        for filename, commands in commands_dict.items():
            for command in commands:
                writer.writerow([filename, command])
        
    return status_data

def sanitize_input_text_file(file, report_folder):
    """Preprocess log file by removing command lines."""
    """This function depends on the Vendor command structure."""
    """Change commands pattern according to the vendor of the device."""
    temp_file = report_folder / f"tmp_{file.name}"
    with open(file, "r", encoding='utf8') as logfile, open(temp_file, "w", encoding='utf8') as tmpf:
        all_lines = logfile.read()
        ## The following line is dependent on the vendor of the device.
        commands = sorted(re.findall(r"^> (show .*|stat .*)", all_lines, re.MULTILINE), key=len, reverse=True)
        #commands = sorted(re.findall(r"^[\w\-_\.]*> (show .*|stat .*)", all_lines, re.MULTILINE), key=len, reverse=True) ############### Use this if prompt is there
        #logging.debug(f"Number of commands: {len(commands)}\n")
        #logging.debug(f"Command list: {commands}\n")
        logfile.seek(0)
        flag_change = False
        for line in logfile:
            for command in commands:
                if line.startswith(command):
                    # Remove the command from the line
                    remainder = line[len(command):]
                    # If the remainder (ignoring whitespace) is non-empty, write it
                    if remainder.strip():
                        tmpf.write(remainder)
                    flag_change = True
                    logging.debug(f"****{command}**** is in the line below:\n{line}")
                    break
            else:
                tmpf.write(line)
                #logging.debug(f"***Line: {line}")
                #logging.debug(f"***No command is in {logfile}")
    if flag_change:
        temp_file.replace(file)
    else:
        temp_file.unlink()
    return commands

def parse_citrix_cmd_output(file):
    """Parse log file and extract device data."""
    device_data = CitrixData( ## max value won't be in all the devices; thus it won't be created here.
        cpu_util={'Current':''}, memory_util={'Current':''}, mgmt_cpu_util={'Current':''}, fans={'health':{'RPM': 0, 'State': ''}}, powers={'health':''}, temperature={'health':''},
        ha={'health':'', 'state':{'State':'', 'Duration':''}, 'sync':'', 'prop':''},
        slb_count={'vs_count':{'UP':0, 'DOWN':0, 'Out Of Service':0}, 'svc_count':{'UP':0, 'DOWN':0, 'Out Of Service':0}, 'svr_count':{'ENABLED':0, 'DISABLED':0}},
        ssl_certs={}, #'valid_cert_name':'dates_until_expiration'
        tcp_cc_client={'Current':0}, tcp_cc_server={'Current':0}, tcp_cps={'Current':0},
        http_rps={'Current':0},
        ssl_sps={'Current':0}, ssl_tps={'Current':0}, key_eps_rsa2048={'Current':0}, key_eps_ecdhe256={'Current':0},
        throughput={'Current':0}
        )
    #logging.debug(f"device_data structure: {device_data}\n")
    with open(file, "r", encoding='utf8') as logfile:
        flag_throughput = False
        flag_frontend = False
        slb_count_key = {} ## This will have only one key and one value at a time
        for line in logfile:
            match line.split():
                case ['Platform:', platform, *rest]:
                    if device_data.model: ## This only works in MPX platform # Platform: NSMPX-8900 8*CPU+32GB+4*F1X+6*E1K+1*E1K+1*COL 8955 30010
                        device_data.model = re.sub(r'(^\D+)\d+(\D*.*$)', r'\1' + device_data.model + r'\2', platform)
                    else:
                        device_data.model = platform
                case ['Model', 'Number', 'ID:', license_model]:
                    if device_data.model: ## This only works in MPX platform # Platform: NSMPX-8900 8*CPU+32GB+4*F1X+6*E1K+1*E1K+1*COL 8955 30010
                        device_data.model = re.sub(r"(^\D+)\d+(\D*.*$)", lambda m: f"{m.group(1)}{license_model}{m.group(2)}", device_data.model)
                    else:
                        device_data.model = license_model
                    #logging.debug(f"model now: {device_data.model}")
                case ['Serial', 'no:', serial]:
                    device_data.serial = serial
                case ['NetScaler', major, 'Build', minor, *rest]: ## NetScaler NS13.1: Build 49.13.nc,
                    device_data.os_version = major.replace(':','-') + minor.strip('.nc,')
                case ['set', 'ns', 'hostName', hostname]:
                    device_data.hostname = hostname
                case ['Up', 'since(Local)', *uptimeList]:
                    ## join a list of strings ## Up since(Local) Fri Aug 16 11:17:25 2024
                    device_data.uptime = " ".join(uptimeList)
                case ['InUse', 'Memory', '(%)', memory]:
                    device_data.memory_util['Current'] = memory ## f"{memory:.2f}"
                case ['Packet', 'CPU', 'usage', '(%)', cpu]:
                    device_data.cpu_util['Current'] = cpu
                case ['Management', 'CPU', 'usage', '(%)', cpu]:
                    device_data.mgmt_cpu_util['Current'] = cpu
                case [fan_type, 'Fan', fan_id, 'Speed', '(RPM)', rpm]:
                    device_data.fans[f"{fan_type} fan {fan_id}"] = rpm ## fan_type: System or CPU
                    if device_data.fans['health']['State'] in {'', 'Normal'}:
                        if int(rpm) < 500:
                            device_data.fans['health']['State'] = 'Not ok'
                            device_data.fans['health']['RPM'] = rpm
                        else:
                            device_data.fans['health']['State'] = 'Normal'
                            device_data.fans['health']['RPM'] = max(int(device_data.fans['health']['RPM']), int(rpm))
                case ['System', 'Fan', 'Speed', '(RPM)', rpm]:
                    device_data.fans[f"System fan"] = rpm
                    if device_data.fans['health']['State'] in {'', 'Normal'}:
                        if int(rpm) < 500:
                            device_data.fans['health']['State'] = 'Not ok'
                            device_data.fans['health']['RPM'] = rpm
                        else:
                            device_data.fans['health']['State'] = 'Normal'
                            device_data.fans['health']['RPM'] = max(int(device_data.fans['health']['RPM']), int(rpm))
                case ['Internal', 'Temperature', '(Celsius)', temperature]:
                    device_data.temperature['degree'] = temperature
                    if int(temperature) < 50:
                        device_data.temperature['health'] = 'Normal'
                case ['Power', 'supply', pid, 'status', status]:  ## only considers 'NORMAL' and 'FAILED'; Not considering 'NOT SUPPORTED'
                    device_data.powers[f"Power {pid}"] = status
                    if device_data.powers['health'] != 'FAILED':
                        device_data.powers['health'] = status
                case ['set', 'ns', 'config', '-IPAddress', ns_ip, *rest]:
                    device_data.ns_ip = ns_ip
                case ['add', 'route', '0.0.0.0', '0.0.0.0', gw]:
                    device_data.gateway = gw
                case ['Node', 'State:', health] if not device_data.ha['health']:
                    device_data.ha['health'] = health
                case ['Master', 'State:', master_type] if not device_data.ha['state']['State']:
                    device_data.ha['state']['State'] = master_type
                case ['Node', 'in', 'this', 'Master', 'State', 'for:', time, '(days:hrs:min:sec)']:
                    days, hours, minutes, seconds = time.split(":")
                    device_data.ha['state']['Duration'] = f"{days} days, {hours} hours, {minutes} minutes, {seconds} seconds"
                case ['Sync', 'State:', state] if not device_data.ha['sync']:
                    device_data.ha['sync'] = state  ## ENABLED if primary, SUCCESS if secondary, no field if there's no HA pair
                case ['Propagation:', state] if not device_data.ha['prop']:
                    device_data.ha['prop'] = state  ## ENABLED/DISABLED, no field if there's no HA pair
                case ['#', 'SSL', 'cards', 'UP', count]:
                    device_data.ssl_cards = int(count)
                case [num, 'Name:', cert_name]:
                    pass ## No need to use a flag unless logic to get device_data.ssl_certs changes.
                case ['Status:', 'Valid,', 'Days', 'to', count]:
                    count = int(count.split(":")[1]) ## expiration:2155
                    if count < 60:  ## This needs to be checked #####################
                        device_data.ssl_certs[cert_name] = count
                case [prompt, 'stat', 'servi', '|', 'grep', state, '-c'] if '>' in prompt: ## 
                    if state == 'OUT':
                        slb_count_key['svc_count'] = 'Out Of Service'
                    else:
                        slb_count_key['svc_count'] = state
                case [prompt, 'stat', 'service', '|', 'grep', state, '-c'] if '>' in prompt: ## If the commands are cut short, this might not work
                    if state == 'OUT':
                        slb_count_key['svc_count'] = 'Out Of Service'
                    else:
                        slb_count_key['svc_count'] = state
                case [prompt, 'stat', 'lb', 'vs', '|', 'grep', state, '-c'] if '>' in prompt: ##
                    if state == 'OUT':
                        slb_count_key['vs_count'] = 'Out Of Service'
                    else:
                        slb_count_key['vs_count'] = state
                case [prompt, 'show', 'server', '-summary', '|', 'grep', 'E', '-c'] if '>' in prompt: ##
                    slb_count_key['svr_count'] = 'ENABLED'
                case [prompt, 'show', 'server', '-summary', '|', 'grep', 'DIS', '-c'] if '>' in prompt: ##
                    slb_count_key['svr_count'] = 'DISABLED'
                case [num] if num.isdigit() and slb_count_key:
                    key, value = slb_count_key.popitem()  ## This will have only one key and one value at a time
                    device_data.slb_count[key][value] = int(num)
                case ['Throughput', 'Statistics','Rate', '(/s)', 'Total']:
                    flag_throughput = True
                case ['Megabits', 'received', throughput, rest] if flag_throughput:
                    device_data.throughput['Current'] = int(throughput)
                    flag_throughput = False
                case ['Established', 'client', 'connections', client_tcp_cc, server_tcp_cc]:
                    device_data.tcp_cc_client['Current'] = int(client_tcp_cc)
                    device_data.tcp_cc_server['Current'] = int(server_tcp_cc)
                case ['Opened', 'client', 'connections(Rate)', client_tcp_cps, server_tcp_cps]:
                    device_data.tcp_cps['Current'] = int(server_tcp_cps)
                case ['Total', 'requests', http_rps, rest]:
                    device_data.http_rps['Current'] = int(http_rps)
                case ['SSL', 'sessions', '(Rate)', ssl_sps]:
                    device_data.ssl_sps['Current'] = int(ssl_sps)
                case ['SSL', 'transactions', ssl_tps, rest]:
                    device_data.ssl_tps['Current'] = int(ssl_tps)
                case ['Front', 'End']:
                    flag_frontend = True
                case ['RSA', '2048-bit', 'key', 'exchanges', key_eps_rsa2048, rest] if flag_frontend:
                    device_data.key_eps_rsa2048['Current'] = int(key_eps_rsa2048)
                    #flag_frontend = False ## Should be enabled at the last component (ECDHE)
                case ['ECDHE', '256', 'curve', 'key', 'exchanges', key_eps_ecdhe256, rest] if flag_frontend:
                    device_data.key_eps_ecdhe256['Current'] = int(key_eps_ecdhe256)
                    flag_frontend = False
                case [weekday, month, date, time, 'KST', year]: ## Mon Feb 17 15:01:14 KST 2025
                    device_data.date_time = parse_and_format_datetime(line.strip())
                case ['License', 'Type:', *license_type]:
                    device_data.license_type = " ".join(license_type)
        #logging.debug(f"Extracted data for {file.name}: {device_data}")
        device_data.log_file_name = file.name
        return device_data

def parse_and_format_datetime(date_string):
    """
    Parse the custom date string and return a formatted string in Korean style.
    Args:
        date_string (str): The date string to parse.
    Returns:
        str: The formatted date and time string.
    """
    # Remove the timezone ("KST") for initial parsing
    date_string_no_tz = date_string.replace("KST", "").strip()
    # Parse the date and time (Input format needs to be specified and changed if input data format changed.)
    dt = datetime.strptime(date_string_no_tz, "%a %b %d %H:%M:%S  %Y")
    # Add the timezone info (KST = UTC+9)
    dt_with_tz = dt.replace(tzinfo=ZoneInfo("Asia/Seoul"))
    
    # Format the datetime into the desired string format
    formatted_datetime = dt_with_tz.strftime("%Y년 %m월 %d일 %H시")
    
    return formatted_datetime

def write_device_data_to_json(report_folder, status_data):
    """Save device data to JSON and CSV."""
    data_json_file = report_folder / f"data_{datetime.today().strftime('%Y%m%d')}.json"
    data_csv_file = report_folder / f"data_{datetime.today().strftime('%Y%m%d')}.csv"

    # Convert the dictionary to a list of dictionaries
    data_list = [{"key": key, **value} for key, value in status_data.items()]
    # Sort the data by the 'hostname' field
    sorted_data = sorted(data_list, key=lambda x: x.get("hostname", "").lower())

    # Save to JSON file
    with open(data_json_file, 'w') as json_f:
        json.dump(sorted_data, json_f, indent=4)
    logging.info(f"Output data file: {data_json_file}")

    with open(data_csv_file, 'w', newline='') as csv_f:
        writer = csv.DictWriter(csv_f, fieldnames=sorted_data[0].keys())
        writer.writeheader()
        writer.writerows(sorted_data)
        #csv_writer = csv.writer(csv_f, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        #csv_writer.writerow(next(iter(sorted_data.values())).keys())
        #for device in sorted_data.values():
        #    csv_writer.writerow(device.values())
    logging.info(f"Output data file: {data_csv_file}")

def validate_hostnames(data, excel_file, inven_sheet, sheets_to_update):
    """
    Fills an Excel sheet with data from a nested dictionary using a mapping between Excel columns and dictionary keys.
    Args:
        data (dict): The nested data dictionary to fill into the Excel file.
        excel_file (str): Path to the Excel file.
        inven_sheet (str): The name of the sheet to check the data against.
        sheets_to_update (list): The name of the sheet to update.
    Returns:
        None
    """
    # Load the workbook and select the sheet
    wb = load_workbook(excel_file)
    ws_inventory = wb[inven_sheet]
    logging.debug(f"Validate Input worksheet: {ws_inventory.title}")
    # Get serial number - hostname, row, hostname column from the inventory worksheet
    records = get_records(ws_inventory)
    # Check inventory sheet against data dictionary if any hostname is different
    # Then copy the sheet and update hostnames that are different on the copied sheet
    updated_hostnames = {}
    updated_inven_ws = wb.copy_worksheet(wb[inven_sheet])
    updated_inven_ws.title = f"{inven_sheet}_new"
    for record in records.values():
        try:
            device_hostname = get_record_by_serial(data, record['serial'])['hostname']
        except KeyError:
            logging.warning(f"Input data with this serial {record['serial']} does NOT exist!")
        if record['hostname'] != device_hostname:
            logging.debug(f"Row number : {record['row']} -- Column number : {record['host_column']}")
            updated_inven_ws[record['row']][record['host_column']-1].value = device_hostname
            updated_inven_ws[record['row']][record['host_column']-1].fill = WARN_FILL
            logging.warning(f"Hostname changed: Serial {record['serial']} - Hostname {record['hostname']} to Hostname now {device_hostname}")
            updated_hostnames[record['hostname']] = device_hostname
    # Update the summary sheet as well.
    if updated_hostnames:
        wb.save(excel_file)
        wb.close()
        for sheet in sheets_to_update:
            find_and_replace(excel_file, sheet, updated_hostnames)
    else:
        wb.close()

def get_records(ws, start_row=2, hostname_col=2, location_col=3, serial_col=4):
    """
    Iterates through rows starting at start_row until an empty cell is encountered in the serial column.
    Builds a dictionary where each key is the serial value, and the value is a dictionary containing:
      - 'serial': serial cell's value
      - 'hostname': hostname cell's value
      - 'host_column': the column index of the hostname cell
      - 'row': the row number
      - 'location': location cell's value
    """
    records = {}
    row = start_row
    while True:
        host = ws.cell(row=row, column=hostname_col)
        location = ws.cell(row=row, column=location_col)
        serial = ws.cell(row=row, column=serial_col)
        
        # Stop when the serial cell is empty (you could also check host.value if that's more reliable)
        if serial.value is None:
            break
        
        records[serial.value] = {
            'serial': serial.value,
            'hostname': host.value,
            'host_column': host.column,  # Note: this may be a number or a letter depending on your openpyxl version.
            'row': host.row,
            'location': location.value
        }
        row += 1
    return records

def find_and_replace(excel_file, sheet_name, replacements):
    # Load the workbook and select the specified sheet
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    
    for row in ws.iter_rows():
        for cell in row:
            # Check if the current cell's value is one of the keys to replace
            if cell.value in replacements:
                cell.value = replacements[cell.value]
    # Save the changes back to the workbook
    wb.save(excel_file)
    wb.close()

def get_record_by_serial(data_dict, serial_value):
    for record in data_dict.values():
        if record.get("serial") == serial_value:
            return record
    logging.warning(f"There's no record of the serial {serial_value}.")
    logging.debug(f"Take a look at record: {record}.")
    return None

def fill_excel_with_nested_data(data, excel_file, sheet_name, mapping):
    """
    Fills an Excel sheet with data from a nested dictionary using a mapping between Excel columns and dictionary keys.
    Args:
        data (dict): The nested data dictionary to fill into the Excel file.
        excel_file (str): Path to the Excel file.
        sheet_name (str): The name of the sheet to update.
        mapping (dict): A mapping of nested dictionary keys to Excel column headings.
    Returns:
        None
    """
    # Load the workbook and select the sheet
    wb = load_workbook(excel_file)
    sheet = wb[sheet_name]

    # Get the hostnames and locations from the 2nd and 3rd row of the sheet
    hostnames = [cell.value for cell in sheet[2][2:]]
    #locations = [cell.value for cell in sheet[3][2:]]
    logging.debug(f"\t{len(hostnames)} hostnames from the template:\n{hostnames}")
    #logging.debug(f"\t{len(locations)} Locations from the template:\n{locations}")
    # Get the row headings from the second column of the sheet
    row_headings = [cell.value for cell in sheet['B'][1:]]
    #logging.debug(f"{len(row_headings)} row headings from the template: {row_headings}")
    # Create a reverse mapping for easier lookup (Excel heading -> dictionary key)
    #reverse_mapping = {v: k for k, v in mapping.items()}

    # Write data into the Excel sheet
    for item in data.values():
        # First data starts in 3rd column and the column index will be found with the hostname of data item.
        col_idx = 3 + hostnames.index(item['hostname'])
        #logging.debug(f"{item} from data dictionary")
        # Enumeration starting number will be the row number for the first data
        for row_idx, heading in enumerate(row_headings, start=2):
            # Check if the heading exists in the reverse mapping
            #logging.debug(f"The row heading from {sheet_name}: {row_idx}.{heading}")
            if heading in mapping:
                # Get the corresponding dictionary key (can be nested)
                dict_key = mapping[heading]
                # Support nested dictionary keys
                keys = dict_key.split('.')  # Split nested keys using '.'
                value = item
                for key in keys:
                    value = value.get(key, None)  # Safely access nested keys
                # Write the data into the cell
                formatted_value = format_value(value)
                #logging.debug(f"Formated value {formatted_value} going into cell({row_idx},{col_idx})")
                sheet.cell(row=row_idx, column=col_idx, value=formatted_value)
    # Save the workbook
    wb.save(excel_file)
    wb.close()
    logging.info(f"Data has been successfully written to {excel_file} in sheet {sheet_name}.")

def process_excel_output_files(data, excel_file, sample_sheet, summary_sheet):
    """
    Fills an Excel sheet with data from a dictionary and a summary sheet.
    Args:
        data (dict): The nested data dictionary to fill into the Excel file.
        excel_file (str): Path to the Excel file.
        sample_sheet (str): The name of sample sheet to copy
        summary_sheet (str): The name of data sheet to copy
    Returns:
        None
    """
    # Load the workbook and select the sheet
    wb = load_workbook(excel_file)
    ws_summary = wb[summary_sheet]
    hostnames = {
        host.value: {'locations': cell.value, 'column': cell.column}
        for host, cell in zip(ws_summary[2][2:], ws_summary[3][2:])
    }
    #logging.debug(f"Hostnames & locations: {hostnames}")
    #logging.debug(f"Data dictionary: {data}")
    # Write data into the Excel sheet
    for idx, hostname in enumerate(hostnames.keys(), start=1):
        if not data.get(hostname):
            logging.debug(f"{idx}. {hostname} doesn't exist in data dictionary.############")
            continue
        ws_device_name = f"{idx}. {hostname}"
        copied_ws = wb.copy_worksheet(wb[sample_sheet])
        copied_ws.title = ws_device_name
        #logging.debug(f"{idx}. {hostname} worksheet: {copied_ws.title}")
        copied_ws['A1'].value = f"정기점검 Check List ({idx})"
        copied_ws["B3"].value = data[hostname]['model']
        copied_ws["D3"].value = hostnames[hostname]['locations']
        copied_ws["F3"].value = data[hostname]['ns_ip']
        copied_ws["B4"].value = hostname
        copied_ws["D4"].value = data[hostname]['serial']
        copied_ws["F4"].value = data[hostname]['os_version']

        # Loop over rows 8 through 30 in the host column
        # and paste them to rows 15 through 37 in column E (which is column 5)
        for i, row in enumerate(range(8, 31)):
            # Get the value from the host column in the summary sheet
            value = ws_summary.cell(row=row, column=hostnames[hostname]['column']).value
            #logging.debug(f"Summary sheet {ws_summary.title} value: {value}")
            # Write that value into column E in the target sheet
            copied_ws.cell(row=14 + i, column=5, value=value)
    # Save the workbook
    wb.save(excel_file)
    logging.info(f"Data has been successfully written to {excel_file} in each device sheet.")

def format_value(value):
    """
    Format the value for writing into the Excel sheet.
    - Lists are joined with commas.
    - Dictionaries are converted to 'key: value' pairs separated by a newline. (To be visible in Excel, the cell has to have "Alignment(wrap_text=True)")
    - Other types are converted to strings.
    """
    if isinstance(value, list):
        return ", ".join(map(str, value))
    elif isinstance(value, dict):
        return "\n".join(f"{k}: {v}" for k, v in value.items())
    else:
        return str(value)

## From here on, the functions will be used for processing Citrix report csv files as input
def read_csv_files_by_hostname(base_directory):
    """
    Reads all CSV files in each subdirectory (each subdirectory is assumed to be a hostname).
    Returns a dictionary where each key is the hostname (subdirectory name) and the value is a list 
    containing the contents of each CSV file (each CSV content is represented as a list of rows).
    """
    base_path = Path(base_directory)
    hostname_data = {}

    # Iterate over each subdirectory in the base directory
    for host_dir in base_path.iterdir():
        if host_dir.is_dir():
            csv_contents = []
            # Iterate over CSV files in the subdirectory
            for csv_file in host_dir.glob("*.csv"):
                with csv_file.open(newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    file_data = [row for row in reader]
                    csv_contents.append(file_data)
            if csv_contents:
                hostname_data[host_dir.name] = csv_contents
    return hostname_data

def compute_max_values(hostname_data):
    """
    Processes the CSV data for each hostname to find, for each metric column (from the CSV header),
    the maximum numeric value and the corresponding time (assumed to be in the first column).
    
    Returns a dictionary of the form:
    {
      'hostname1': {
         'Metric Heading 1': {'MAX': max_value, 'Time': corresponding_time},
         'Metric Heading 2': {'MAX': max_value, 'Time': corresponding_time},
         ...
      },
      'hostname2': { ... },
      ...
    }
    """
    result = {}
    for hostname, csv_files in hostname_data.items():
        max_dict = {}  # This will store our metric maximums for the current hostname
        
        # Process each CSV file for this hostname
        for csv_data in csv_files:
            if not csv_data or len(csv_data) < 2:
                continue  # Skip if no header or no data rows
            header = csv_data[0]  # The header row; assume first column is "Time"
            # Loop through each data row (skip header)
            for row in csv_data[1:]:
                if not row:
                    continue
                # Assume the first column is the timestamp
                time_val = row[0]
                # Process each metric column (starting from index 1)
                for i in range(1, len(header)):
                    heading = header[i]
                    try:
                        # Convert the cell value to a float
                        value = float(row[i])
                    except (ValueError, TypeError):
                        continue  # Skip if the conversion fails
                    # If we haven't seen this metric yet, add it
                    if heading not in max_dict:
                        max_dict[heading] = {'MAX': value, 'Time': time_val}
                    else:
                        # Update if we find a higher value
                        if value > max_dict[heading]['MAX']:
                            max_dict[heading]['MAX'] = value
                            max_dict[heading]['Time'] = time_val
        result[hostname] = max_dict
    return result

def dict_get_nested(data, outer_key, inner_key):
    outer_value = data.get(outer_key)
    if outer_value is None:
        logging.warning(f"Key '{outer_key}' is missing in the dictionary.")
        return None
    inner_value = outer_value.get(inner_key)
    if inner_value is None:
        logging.warning(f"Key '{inner_key}' is missing under '{outer_key}'.")
        return None
    return inner_value

def highlight_flagged_columns(
    filename,
    header_col_letter="A",   # the left-most column that contains row labels
    sheetname="data",
    flag_value="State: Primary",
    fill_color='e4fbd5'
):
    """
    Loads an Excel workbook and processes the data sheet.  
    For each column (except the header column) that contains at least one cell 
    (from row 2 down) whose text includes 'State: Primary', highlight that 
    column by filling the cells (starting at row 2) only if the corresponding 
    cell in the header column (left-most column) is not empty.
    
    Parameters:
        filename (str): Path to the Excel workbook.
        header_col_letter (str): Letter of the header column (default "A").
        sheetname (str): Worksheet name to process (default "data").
        fill_color (str): Hex fill color to use (default "57F78E").
    """
    # Define the fill style for highlighting.
    primary_fill = PatternFill(fill_type="solid", fgColor=fill_color)
    
    # Load the workbook and worksheet.
    wb = load_workbook(filename)
    ws = wb[sheetname]
    
    # Convert the header column letter to a 1-indexed column number.
    header_col_index = column_index_from_string(header_col_letter)
    
    # Identify columns to highlight.
    # We scan each column (starting from row 2, because row 1 is the header row).
    # We skip the header column itself because that column is only for row labels.
    cols_to_highlight = set()
    for col_cells in ws.iter_cols(min_row=2, max_row=ws.max_row, values_only=False):
        # Skip the header column (e.g. column A).
        if col_cells[0].column == header_col_index:
            continue
        for cell in col_cells:
            if cell.value and flag_value in str(cell.value):
                cols_to_highlight.add(cell.column)  # cell.column is 1-indexed.
                break  # No need to check further cells in this column.
    
    # For every column that qualifies, highlight the cell (from row 2 onward)
    # only if the corresponding cell in the header column (e.g. A) is not empty.
    for col in cols_to_highlight:
        for row in range(2, ws.max_row + 1):
            header_cell = ws.cell(row=row, column=header_col_index)
            target_cell = ws.cell(row=row, column=col)
            # Only apply the fill if both the header cell and the target cell have a value.
            #if header_cell.value not in (None, "") and target_cell.value not in (None, ""):
            # Only apply the fill if the header cell has a value.
            if header_cell.value not in (None, ""):
                target_cell.fill = primary_fill
    wb.save(filename)


if __name__ == "__main__":
    ##############################################
    # Initialize input and output path
    input_folder, report_folder = initialize_paths()
    # Setup logger for debugging & critical errors
    logging = Logger.configure_logging(report_folder)
    logging.info(f"Script started processing...")
    logging.info(f"Input folder: {input_folder}")
    ##############################################
    # Prepare input data
    cmd_csv_file = report_folder / f"commands_{datetime.today().strftime('%Y%m%d')}.csv"
    status_data = load_and_prepare_input_files(input_folder, report_folder, cmd_csv_file)
    # Read in csv input data
    data_from_report = read_csv_files_by_hostname(input_folder)
    max_values = compute_max_values(data_from_report)
    #logging.debug(f"MAXIMUM Values Dictionary : {max_values}")
    for hostname, max_value_for_hostname in max_values.items():
        #logging.debug(f"Maximum data for {hostname}: {max_value_for_hostname}")
        logging.debug(f"Processing maximum data for {hostname}")
        status_data[hostname]['cpu_util']['MAX'] = dict_get_nested(max_value_for_hostname,'Packet CPU usage (%)', 'MAX')
        status_data[hostname]['memory_util']['MAX'] = dict_get_nested(max_value_for_hostname,'InUse Memory (%)', 'MAX')
        status_data[hostname]['tcp_cc_client']['MAX'] = dict_get_nested(max_value_for_hostname,'Established client connections', 'MAX')
        status_data[hostname]['tcp_cc_server']['MAX'] = dict_get_nested(max_value_for_hostname,'Established server connections', 'MAX')
        status_data[hostname]['http_rps']['MAX'] = dict_get_nested(max_value_for_hostname,'Total requests (Rate)', 'MAX')
        status_data[hostname]['ssl_sps']['MAX'] = dict_get_nested(max_value_for_hostname,'SSL sessions  (Rate)', 'MAX')
        status_data[hostname]['key_eps_rsa2048']['MAX'] = dict_get_nested(max_value_for_hostname,'RSA 2048-bit key exchanges (Rate)', 'MAX')
        status_data[hostname]['key_eps_ecdhe256']['MAX'] = dict_get_nested(max_value_for_hostname,'ECDHE 256 curve key exchanges (Rate)', 'MAX')
        status_data[hostname]['throughput']['MAX'] = dict_get_nested(max_value_for_hostname,'Megabits received (Rate)', 'MAX')
    logging.info(f"Input file processing is completed..")
    ##############################################
    # Prepare output file
    # Mapping of dictionary keys to Excel row headings
    # Mapping of nested dictionary keys to Excel row headings e.g. 'intf_errors.error_types.input_errors': 'Intf Rx Errors'
    templates = {
        'default':  {'file_name': 'template_citrix.xlsx', 'data_sheet': '점검결과_Review', 
                        'row_mapping': {
                            'Hostname': 'hostname', 'Model': 'model', 'Serial Number': 'serial', 
                            'NSIP': 'ns_ip', 'Version': 'os_version', 'System Uptime': 'uptime', 
                            'CPU 사용률': 'cpu_util', 'Memory 사용률': 'memory_util', 
                            'Power Supply State': 'powers.health', 'Fan State': 'fans.health', 
                            'Interface State': 'intf_errors', 'Node State': 'ha.health', 
                            'Master State': 'ha.state', 'Sync State': 'ha.sync', 'Propagation': 'ha.prop', 
                            'Virtual Server': 'slb_count.vs_count', 'Service': 'slb_count.svc_count', 
                            'Server': 'slb_count.svr_count', 'SSL Cards State': 'ssl_cards', 
                            'Certificate DaytoExpire State': 'ssl_certs', 'System Throughput(Mbps)': 'throughput', 
                            'TCP Eastablished Client Conn': 'tcp_cc_client', 'TCP Eastablished Server Conn': 'tcp_cc_server', 
                            'HTTP Requests Rate(/s)': 'http_rps', 'SSL Session Rate(/s)': 'ssl_sps', 
                            'Key Exchange Rate(/s) / RSA2048': 'key_eps_rsa2048', 'Key Exchange Rate(/s) / ECDHE 256 Curve': 'key_eps_ecdhe256', 
                            'Log State': 'log'
                        }},
        '1ws':      {'file_name': 'template_citrix_1ws.xlsx', 'data_sheet': '점검결과_Review',
                        'row_mapping': {
                            'HostName': 'hostname', 'Model': 'model', 'Hardware Serial': 'serial', 
                            'NSIP': 'ns_ip', 'Version': 'os_version', 'System Uptime': 'uptime', 'License': 'license_type',
                            'CPU 부하': 'cpu_util', 'MGMT CPU 부하': 'mgmt_cpu_util','Memory 사용률': 'memory_util', 
                            'Power': 'powers.health', 'Power Supply 이중화': 'powers.health', 'Fan': 'fans.health', 
                            'Interface 상태': 'intf_errors', 'HA 상태': 'ha.health', 
                            'Configuration Sync 상태': 'ha.state.State', 'Sync State': 'ha.sync', 'Propagation': 'ha.prop', 
                            'Virtual-Server Total Count': 'slb_count.vs_count.total', 'Service Total Count': 'slb_count.svc_count.total', 
                            'Real Server Total Count': 'slb_count.svr_count.total', 'Module or Card': 'ssl_cards', 
                            'Certificate DaytoExpire State': 'ssl_certs', 'System Throughput(Mbps)': 'throughput', 
                            'TCP Eastablished Client Conn': 'tcp_cc_client', 'TCP Eastablished Server Conn': 'tcp_cc_server', 
                            'HTTP Requests Rate(/s)': 'http_rps', 'SSL Session Rate(/s)': 'ssl_sps', 
                            'Key Exchange Rate(/s) / RSA2048': 'key_eps_rsa2048', 'Key Exchange Rate(/s) / ECDHE 256 Curve': 'key_eps_ecdhe256', 
                            'Log State': 'log'
                        }}
    }
    template_file = Path("./template_citrix.xlsx")
    logging.info(f"Template file: {template_file}")
    report_file = report_folder / f"Report_citrix_{datetime.today().strftime('%Y%m%d')}.xlsx"
    logging.info(f"Report output file: {report_file}")
    # Load a template file and copy a new workbook from the template
    template_wb = load_workbook(template_file, read_only=False, data_only=False)
    template_wb.save(report_file)
    template_wb.close()
    inven_sheet = 'Inven'
    data_sheet = '점검결과_Review'  ## Worksheet with data for engineers
    sheets_to_update_hostnames = [data_sheet, '점검결과_Summary'] ## Sheets in this list will be changed, should any hostname changed.
    # Validate hostname-serial number pairs are all correct
    # If they aren't the same, the sheets will be updated
    logging.info(f"Start comparing hostnames of input file against hostnames of an inventory sheet")
    validate_hostnames(status_data, report_file, inven_sheet, sheets_to_update_hostnames)
    # Update a summary data sheet
    logging.info(f"Now the summary data sheet is being updated.....")
    fill_excel_with_nested_data(status_data, report_file, data_sheet, templates['default']['row_mapping'])
    # Color the summary data sheet to become more readable
    highlight_flagged_columns(report_file, header_col_letter='B', sheetname=data_sheet, fill_color=HA_PRIMARY_COLOR)
    #logging.debug(f"Data dictionary:\n{status_data}")
    # Copy a sample sheet and dump data onto a copied sheet for each device
    logging.info(f"Each device data is being written on its own sheet.....")
    process_excel_output_files(status_data, report_file, 'sample', data_sheet)
    #output_wb.remove(output_wb["sample"])
    #output_wb.save(reportFile)

    logging.info(f"Processing complete.")
    logging.info(f"Report folder: {report_folder}")
